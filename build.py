#!/usr/bin/env python3
"""
build.py — Đọc data/latest.xlsx, xử lý dữ liệu, và build ra index.html
từ template.html.

Cách chạy thủ công (nếu cần test ở máy local):
    pip install openpyxl
    python3 build.py

GitHub Actions sẽ tự động chạy script này mỗi khi data/latest.xlsx
được cập nhật (xem .github/workflows/build.yml).
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Thiếu thư viện openpyxl. Chạy: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# CẤU HÌNH ĐƯỜNG DẪN
# ---------------------------------------------------------------------------
REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(REPO_ROOT, "data", "latest.xlsx")
TEMPLATE_FILE = os.path.join(REPO_ROOT, "template.html")
OUTPUT_FILE = os.path.join(REPO_ROOT, "index.html")

# ---------------------------------------------------------------------------
# BẢNG ÁNH XẠ CHI NHÁNH -> VÙNG
# Nếu công ty mở chi nhánh mới, chỉ cần thêm 1 dòng vào đây rồi chạy lại.
# Chi nhánh nào không có trong bảng này sẽ hiện "Chưa gán vùng" trên dashboard.
# ---------------------------------------------------------------------------
REGION_MAPPING = {
    "Scots English Kim Giang": "Vùng 1",
    "Scots English Thanh Hóa": "Vùng 1",
    "Scots English Lam Sơn": "Vùng 1",
    "Scots English Nguyễn Xiển": "Vùng 1",
    "Scots English Linh Đàm": "Vùng 1",
    "Scots English Tây Hồ": "Vùng 1",
    "Scots English Nguyễn Tuân": "Vùng 1",
    "Scots English Hoàng Đạo Thúy": "Vùng 1",
    "Scots English Hoàng Quốc Việt": "Vùng 1",
    "Scots English Trung Văn": "Vùng 1",
    "Scots English Hải Dương": "Vùng của Liên",
    "Scots English Vĩnh Phúc": "Vùng của Liên",
    "Scots English Long Biên": "Vùng của Liên",
    "Scots English Vĩnh Phúc 3": "Vùng của Liên",
    "Scots English Phúc Yên": "Vùng của Liên",
    "Scots English Việt Trì": "Vùng của Liên",
    "Scots English Mỹ Đình": "Vùng của Liên",
    "Scots English Vinhomes Gardenia": "Vùng của Liên",
    "Scots English Times City": "Vùng 3",
    "Scots English Văn Khê": "Vùng 3",
    "Scots English An Khánh": "Vùng 3",
    "Scots English Vinhomes Smart City": "Vùng 3",
    "Scots English Vinhomes Smart City 2": "Vùng 3",
    "Scots English Dương Nội": "Vùng 3",
    "Scots English Phạm Văn Đồng": "Vùng 3",
    "Scots English Thái Bình": "Vùng 3",
    "Scots English Từ Sơn": "Vùng 5",
    "Scots English Hải Phòng 2": "Vùng 5",
    "Scots English Hải Phòng": "Vùng 5",
    "Scots English Bắc Ninh": "Vùng 5",
    "Scots English Bắc Ninh 2": "Vùng 5",
    "Scots English Bắc Giang": "Vùng 5",
    "Scots English Sài Đồng": "Vùng 6",
    "Scots English Vinh": "Vùng 6",
    "Scots English Ocean Park": "Vùng 6",
    "Scots English Trường Chinh": "Vùng 6",
    "Scots English Định Công": "Vùng 6",
    "Scots English Đà Nẵng": "Vùng 7",
    "Scots English Đà Nẵng 2": "Vùng 7",
    "Scots English Phan Văn Trị": "Vùng 7",
    "Scots English Celadon - Tân Phú": "Vùng 8",
    "Scots English Phạm Văn Chiêu": "Vùng 8",
    "Scots English Grand Park": "Vùng 8",
}


def region_of(branch):
    return REGION_MAPPING.get(branch, "Chưa gán vùng")


def safe(value, default=0):
    return default if value is None else value


def parse_class_summary_monthly(ws):
    """Sheet 'Class Summary Monthly' -> danh sách dict cho classData."""
    rows = ws.iter_rows(min_row=2, values_only=True)
    out = []
    for row in rows:
        if not row or row[0] is None:
            continue
        branch, program, cls, month, records, att, ab, late, attp, abp, latep = row[:11]
        branch = str(branch).strip()
        out.append({
            "region": region_of(branch),
            "branch": branch,
            "program": program,
            "class_name": cls,
            "month": month,
            "records": safe(records),
            "attendance": safe(att),
            "absence": safe(ab),
            "late": safe(late),
        })
    return out


def parse_student_summary(ws):
    """Sheet 'Student Summary' -> danh sách dict cho studentData."""
    rows = ws.iter_rows(min_row=2, values_only=True)
    out = []
    for row in rows:
        if not row or row[0] is None:
            continue
        branch, program, cls, name, sid, grade, records, att, ab, late, attp, abp, latep = row[:13]
        branch = str(branch).strip()
        out.append({
            "region": region_of(branch),
            "branch": branch,
            "program": program,
            "class_name": cls,
            "name": name,
            "student_id": sid,
            "grade": grade,
            "records": safe(records),
            "attendance": safe(att),
            "absence": safe(ab),
            "late": safe(late),
        })
    return out


def _cell_to_jsonable(value):
    """Chuyển giá trị cell Excel (có thể là datetime/date) thành string/number
    để nhét thẳng vào JSON mà không lỗi."""
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def parse_raw_sheet(ws):
    """Đọc 1 sheet 'thô' (raw) theo ĐÚNG tên cột thật có trong sheet, không cố
    định trước danh sách cột — dùng cho các sheet 'Number of Student' và
    'List of Student' vì tên cột thật từ API có thể khác giữa các lần cập nhật.

    Trả về list các dict với key = tên cột (giữ nguyên as-is trong Excel).
    Nếu sheet có cột 'Branch', tự thêm field 'region' (ánh xạ theo REGION_MAPPING)
    để dashboard có thể lọc theo Vùng.
    """
    if ws is None:
        return []
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    # Bỏ qua cột không có tên (None) — Excel đôi khi có cột thừa/trống ở cuối
    columns = [(i, str(h)) for i, h in enumerate(header) if h is not None and str(h).strip() != ""]

    out = []
    for row in rows_iter:
        if not row or all(v is None for v in row):
            continue
        record = {}
        for i, col_name in columns:
            record[col_name] = _cell_to_jsonable(row[i] if i < len(row) else None)
        if "Branch" in record:
            record["region"] = region_of(str(record["Branch"]).strip())
        out.append(record)
    return out


def main():
    if not os.path.exists(DATA_FILE):
        print(f"Không tìm thấy file dữ liệu: {DATA_FILE}", file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(TEMPLATE_FILE):
        print(f"Không tìm thấy template: {TEMPLATE_FILE}", file=sys.stderr)
        sys.exit(1)

    print(f"Đang đọc {DATA_FILE} ...")
    wb = openpyxl.load_workbook(DATA_FILE, read_only=True, data_only=True)

    required_sheets = ["Class Summary Monthly", "Student Summary"]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        print(
            f"File Excel thiếu sheet bắt buộc: {missing}. "
            f"Các sheet hiện có: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)

    class_data = parse_class_summary_monthly(wb["Class Summary Monthly"])
    student_data = parse_student_summary(wb["Student Summary"])

    number_raw_ws = wb["Number of Student"] if "Number of Student" in wb.sheetnames else None
    list_raw_ws = wb["List of Student"] if "List of Student" in wb.sheetnames else None
    number_raw_data = parse_raw_sheet(number_raw_ws)
    list_raw_data = parse_raw_sheet(list_raw_ws)
    if number_raw_ws is None:
        print("Lưu ý: không có sheet 'Number of Student' -> tab 'Chi tiết' phần Number of Student sẽ trống.")
    if list_raw_ws is None:
        print("Lưu ý: không có sheet 'List of Student' -> tab 'Chi tiết' phần List of Student sẽ trống.")

    unmapped = sorted({
        d["branch"] for d in (class_data + student_data)
        if d["region"] == "Chưa gán vùng"
    })
    if unmapped:
        print("CẢNH BÁO: các chi nhánh sau chưa có trong REGION_MAPPING:")
        for b in unmapped:
            print(f"  - {b}")
        print("Dashboard vẫn build bình thường, các chi nhánh này sẽ hiện 'Chưa gán vùng'.")

    print(
        f"Đã xử lý {len(class_data)} dòng lớp/tháng, {len(student_data)} học viên, "
        f"{len(number_raw_data)} dòng Number of Student (raw), {len(list_raw_data)} dòng List of Student (raw)."
    )

    class_json = json.dumps(class_data, ensure_ascii=False, separators=(",", ":"))
    student_json = json.dumps(student_data, ensure_ascii=False, separators=(",", ":"))
    number_raw_json = json.dumps(number_raw_data, ensure_ascii=False, separators=(",", ":"))
    list_raw_json = json.dumps(list_raw_data, ensure_ascii=False, separators=(",", ":"))
    region_json = json.dumps(REGION_MAPPING, ensure_ascii=False, separators=(",", ":"))

    with open(TEMPLATE_FILE, "r", encoding="utf-8") as f:
        html = f.read()

    required_placeholders = [
        "__CLASS_DATA__", "__STUDENT_DATA__", "__REGION_MAP__",
        "__NUMBER_OF_STUDENT_RAW__", "__LIST_OF_STUDENT_RAW__",
    ]
    missing_ph = [p for p in required_placeholders if p not in html]
    if missing_ph:
        print(
            f"template.html thiếu placeholder: {missing_ph}.",
            file=sys.stderr,
        )
        sys.exit(1)

    html = html.replace("const CLASS_DATA = __CLASS_DATA__;", "const CLASS_DATA = " + class_json + ";")
    html = html.replace("const STUDENT_DATA = __STUDENT_DATA__;", "const STUDENT_DATA = " + student_json + ";")
    html = html.replace("const REGION_MAP = __REGION_MAP__;", "const REGION_MAP = " + region_json + ";")
    html = html.replace("const NUMBER_OF_STUDENT_RAW = __NUMBER_OF_STUDENT_RAW__;", "const NUMBER_OF_STUDENT_RAW = " + number_raw_json + ";")
    html = html.replace("const LIST_OF_STUDENT_RAW = __LIST_OF_STUDENT_RAW__;", "const LIST_OF_STUDENT_RAW = " + list_raw_json + ";")

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(html)

    size_mb = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)
    print(f"Đã tạo {OUTPUT_FILE} ({size_mb:.1f} MB). Xong!")
    if size_mb > 80:
        print(
            f"⚠️  CẢNH BÁO: index.html đang {size_mb:.1f} MB, gần/vượt giới hạn 100MB của GitHub. "
            f"Nếu bước 'git push' bị lỗi vượt giới hạn dung lượng, hãy giảm DETAIL_MONTHS_BACK "
            f"khi chạy scrape.js (điều chỉnh trong workflow_dispatch hoặc biến môi trường)."
        )


if __name__ == "__main__":
    main()
